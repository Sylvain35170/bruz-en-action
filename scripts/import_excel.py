"""
import_excel.py — Met à jour les statuts de data/promesses.json depuis le référentiel Excel
de l'association (input/BEA/referentiel_promesses_bruz.xlsx, feuille "Suivi des promesses").

Ne reconstruit pas promesses.json : met à jour uniquement les promesses existantes,
identifiées par leur `ref` (ex. "M1"), pour ne pas écraser `detail` / `source` (citations
du programme officiel) qui ne viennent pas de l'Excel.

Usage : python3 scripts/import_excel.py
"""

import json
import sys
from pathlib import Path
from datetime import date

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl requis : pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent.parent
INPUT_FILE = ROOT / "input" / "BEA" / "referentiel_promesses_bruz.xlsx"
OUTPUT_FILE = ROOT / "data" / "promesses.json"
SHEET_NAME = "Suivi des promesses"
HEADER_ROW = 3

STATUT_MAP = {
    "à faire": "non_commence",
    "en cours": "en_cours",
    "tenu": "tenu",
    "tenue": "tenu",
    "partiellement tenu": "partiel",
    "partiel": "partiel",
    "abandonné": "abandonne",
    "abandonnée": "abandonne",
    "non évaluable": "inconnu",
    "inconnu": "inconnu",
}


def load_rows():
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb[SHEET_NAME]
    headers = [c.value for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))]
    rows = {}
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if not any(row):
            continue
        record = dict(zip(headers, row))
        ref = str(record.get("N°", "") or "").strip()
        if ref:
            rows[ref] = record
    return rows


def main():
    if not INPUT_FILE.exists():
        print(f"❌ Fichier introuvable : {INPUT_FILE}")
        sys.exit(1)

    excel_rows = load_rows()
    print(f"{len(excel_rows)} promesse(s) lues dans l'Excel.")

    data = json.load(open(OUTPUT_FILE, encoding="utf-8"))
    promesses = data["promesses"]

    seen_refs = set()
    updated = 0
    for promesse in promesses:
        ref = promesse.get("ref")
        excel_row = excel_rows.get(ref)
        if not excel_row:
            continue
        seen_refs.add(ref)

        statut_brut = str(excel_row.get("Statut", "") or "").strip().lower()
        statut_id = STATUT_MAP.get(statut_brut)
        if statut_brut and not statut_id:
            print(f"⚠️  {ref} : statut Excel inconnu « {statut_brut} » → ignoré")
        elif statut_id and statut_id != promesse.get("statut_id"):
            print(f"  {ref} : {promesse.get('statut_id')} → {statut_id}")
            promesse["statut_id"] = statut_id
            updated += 1

        horizon = str(excel_row.get("Horizon\n(année cible)", "") or "").strip()
        if horizon and horizon != promesse.get("horizon"):
            promesse["horizon"] = horizon

        commentaires = str(excel_row.get("Commentaires", "") or "").strip()
        if commentaires:
            promesse["commentaires_asso"] = commentaires

    missing_in_json = set(excel_rows) - seen_refs
    if missing_in_json:
        print(f"⚠️  Réf. présentes dans l'Excel mais absentes de promesses.json : {sorted(missing_in_json)}")

    data["meta"]["last_updated"] = str(date.today())
    data["meta"]["updated_by"] = "import_excel.py"

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✅ {updated} statut(s) mis à jour → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
